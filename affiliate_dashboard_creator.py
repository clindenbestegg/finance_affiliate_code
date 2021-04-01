
""" ------------------------------------ Start of classes -------------------------------------------"""
class Affiliate_Data():
	def __init__(self, name, query_name):
		self.name = name
		self.query_name = query_name
		self.record_data = []

class Record_Data():
	def __init__(self, period, company_name, loan_amt, referral_amt, num_loans_funded):
		self.period = str(period)
		self.company_name = company_name
		self.loan_amt = float(loan_amt)
		self.referral_amt = float(referral_amt)
		self.num_loans_funded = int(num_loans_funded)
		self.cpl = self.calculate_cpl()
		self.month = self.calculate_month()
		self.year = self.calculate_year()
		self.qtr = self.calculate_qtr()

	def calculate_cpl(self):
		try:
			cpl = round(self.referral_amt / self.loan_amt,4)
		except ZeroDivisionError:
			cpl = 0
		return cpl

	def calculate_month(self):
		month = self.period[-2:]
		month = int(month)
		return month

	def calculate_year(self):
		year = self.period[:4]
		year = int(year)
		return year

	def calculate_qtr(self):
		qtr = ""
		if self.month in [1,2,3]:
			qtr = "Q1"
		elif self.month in [4,5,6]:
			qtr = "Q2"
		elif self.month in [7,8,9]:
			qtr = "Q3"
		elif self.month in [10,11,12]:
			qtr = "Q4"
		return qtr

class Output_Row():
	def __init__(self, row_identifier):
		self.row_identifier = row_identifier
		self.total_loans = 0
		self.total_funded_amt = 0
		self.total_referral = 0
		self.cpl = 0

	def add_loans(self, num_loans, funded_amt, referral_amt):
		self.total_loans += num_loans
		self.total_funded_amt += funded_amt
		self.total_referral += referral_amt

	def calc_cpl(self):
		try:
			self.cpl = round(self.total_referral / self.total_funded_amt,4)
		except ZeroDivisionError:
			self.cpl = 0

	def format_empty_data(self):
		if self.total_loans == 0:
			self.total_loans = "---"
		if self.total_funded_amt == 0:
			self.total_funded_amt = "---"
		if self.total_referral == 0:
			self.total_referral = "---"
		if self.cpl == 0:
			self.cpl = "---"

""" ------------------------------------ End of classes ---------------------------------------------"""

""" ------------------------------------ Start of Functions -----------------------------------------"""

""" Data Import Functions"""
def import_data(input_file_path):
	#points to the workbook
	wb = load_workbook(input_file_path + "ITD_affiliate_data_for_dashboard_generation.xlsm")

	#looks at the sheet affiliate list to pull in all affiliates to build sheets for
	ws = wb["Affiliate List"]
	affil_list_max_row = ws.max_row

	#builds an array of affiliates to build lists for
	affil_data = []
	for i in range (2, affil_list_max_row+1):
		if ws.cell(i,3).value == True:
			name = ws.cell(i, 1).value
			query_name = ws.cell(i, 2).value
			affil_data.append(Affiliate_Data(name, query_name))

	#moves to the data tab to begin loading data
	ws = wb["Data"]
	data_list_max_row = ws.max_row

	#parses thru the list to check and see if the record belongs in a affiliate, if so, adds to that affiliate
	for i in range(2, data_list_max_row+1):
		period = ws.cell(i,1).value
		company_name = ws.cell(i,2).value
		loan_amt = ws.cell(i,3).value
		referral_amt = ws.cell(i,4).value
		num_loans = ws.cell(i,5).value

		for j in affil_data:
			if j.query_name == company_name:
				j.record_data.append(Record_Data(period, company_name, loan_amt, referral_amt, num_loans))


	return affil_data

""" End Data Import Functions"""

""" Output Generation Functions"""
def write_output_file(affiliate_class, current_month, current_num_month, current_short_month, current_year, output_file_path):
	file = output_file_path + "Automated Dashboards/" + affiliate_class.name + "_" + current_short_month + "_" + str(current_year) + "_dashboard.xlsx"
	wb = Workbook()
	ws = wb.active

	#sets general borders and gridlines of the output
	ws.sheet_view.showGridLines = False

	#border, fill, alignment, font settings
	border_medium = Side(border_style="medium", color="000000")
	border_thin = Side(border_style="thin", color="000000")

	#alightment styles
	align_center = Alignment(horizontal='center', vertical='center')

	#font styles
	font_title = Font(size = 18, bold = True)
	font_month = Font(size = 16, bold = True)
	font_monthly_results = Font(size = 14, bold = True)
	font_table_titles = Font(size = 12, bold = True)
	font_table_headers = Font(size = 11, bold = True)

	#fill styles
	fill_table_header = PatternFill(start_color = "64bcc3", end_color = "64bcc3", fill_type = "solid")
	fill_table_zebra = PatternFill(start_color = "bce2e5", end_color = "bce2e5", fill_type = "solid")
	fill_table_total = PatternFill(start_color = "d1d3d3", end_color = "d1d3d3", fill_type = "solid")


	#writes and formats the affiliate name
	ws.merge_cells('J3:N4')
	ws['J3'] = affiliate_class.name
	ws['J3'].alignment = align_center
	ws['J3'].font = font_title

	#Large Box Border Settings
		#top and bottoms
	for i in range(2,23):
		ws.cell(row = 5, column = i).border = Border(bottom = border_medium)
		ws.cell(row = 13, column = i).border = Border(top = border_medium, bottom = border_medium)
		ws.cell(row = 25, column = i).border = Border(top = border_medium)

		#top sides
	for i in range(6, 13):
		ws.cell(row = i, column = 2).border = Border(left = border_medium)
		ws.cell(row = i, column = 22).border = Border(right = border_medium)

		#bottom sides
	for i in range(14, 25):
		ws.cell(row = i, column = 2).border = Border(left = border_medium)
		ws.cell(row = i, column = 22).border = Border(right = border_medium)



	#Internal Box Ranges and captions
		#top box
	ws.merge_cells('J6:N7')
	ws.merge_cells('D9:H10')
	ws.merge_cells('D11:H11')
	ws.merge_cells('J9:N10')
	ws.merge_cells('J11:N11')
	ws.merge_cells('P9:T10')
	ws.merge_cells('P11:T11')

	#Internal Top box border settings
		#top and bottom
	for i in range(4, 9):
		ws.cell(row = 9, column = i).border = Border(top = border_thin)
		ws.cell(row = 11, column = i).border = Border(bottom = border_thin)

	for i in range(10, 15):
		ws.cell(row = 9, column = i).border = Border(top = border_thin)
		ws.cell(row = 11, column = i).border = Border(bottom = border_thin)

	for i in range(16, 21):
		ws.cell(row = 9, column = i).border = Border(top = border_thin)
		ws.cell(row = 11, column = i).border = Border(bottom = border_thin)

	for i in range (9,12):
		ws.cell(row = i, column = 3).border = Border(right = border_thin)
		ws.cell(row = i, column = 9).border = Border(left = border_thin, right = border_thin)
		ws.cell(row = i, column = 15).border = Border(left = border_thin, right = border_thin)
		ws.cell(row = i, column = 21).border = Border(left = border_thin)

	#Month title
	ws['J6'].font = font_month
	ws['J6'].alignment = align_center
	ws['J6'].value = f'{current_month} Results'

	#inside top boxes
	ws['D9'].font = font_monthly_results
	ws['D9'].alignment = align_center
	ws['D9'].number_format = "$#,##0.00"
	ws['D11'].value = "Funded Loan Amount"
	ws['D11'].alignment = align_center

	ws['J9'].font = font_monthly_results
	ws['J9'].alignment = align_center
	ws['J9'].number_format = "$#,##0.00"
	ws['J11'].value = "Compensation Amount"
	ws['J11'].alignment = align_center

	ws['P9'].font = font_monthly_results
	ws['P9'].alignment = align_center
	ws['P9'].number_format = "0.00%"
	ws['P11'].value = "Compensation Per Loan"
	ws['P11'].alignment = align_center


	#Bottom Box Creation
		#titles
	ws.merge_cells('F15:H15')
	ws['F15'].alignment = align_center
	ws['F15'].font = font_table_titles
	ws['F15'].value = "YTD by Quarter"

	ws.merge_cells('P15:R15')
	ws['P15'].alignment = align_center
	ws['P15'].font = font_table_titles
	ws['P15'].value = "Last 5 Years YTD"




	#breakdown boxes
		#left box
	ws.merge_cells('C17:D17')
	ws['C17'].border = Border(top = border_medium, bottom = border_medium, left = border_medium)
	ws['D17'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)
	ws['C17'].value = "Quarter"
	ws.merge_cells('C18:D18')
	ws['C18'].border = Border(bottom = border_thin, left = border_medium)
	ws['D18'].border = Border(bottom = border_thin, right = border_medium)
	ws['C18'].value = f'{current_year} - Q1'
	ws.merge_cells('C19:D19')
	ws['C19'].border = Border(bottom = border_thin, left = border_medium)
	ws['D19'].border = Border(bottom = border_thin, right = border_medium)
	ws['C19'].value = f'{current_year} - Q2'
	ws.merge_cells('C20:D20')
	ws['C20'].border = Border(bottom = border_thin, left = border_medium)
	ws['D20'].border = Border(bottom = border_thin, right = border_medium)
	ws['C20'].value = f'{current_year} - Q3'
	ws.merge_cells('C21:D21')
	ws['C21'].border = Border(left = border_medium)
	ws['D21'].border = Border(right = border_medium)
	ws['C21'].value = f'{current_year} - Q4'
	ws.merge_cells('C22:D22')
	ws['C22'].border = Border(top = border_medium, bottom = border_medium, left = border_medium)
	ws['D22'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)
	ws['C22'].value = "Grand Total"

	ws.merge_cells('E17:F17')
	ws['E17'].border = Border(top = border_medium, bottom = border_medium)
	ws['F17'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)
	ws['E17'].value = "Funded Loans"
	ws.merge_cells('E18:F18')
	ws.merge_cells('E19:F19')
	ws.merge_cells('E20:F20')
	ws.merge_cells('E21:F21')
	ws.merge_cells('E22:F22')
	ws['E22'].border = Border(top = border_medium, bottom = border_medium)
	ws['F22'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)


	ws.merge_cells('G17:H17')
	ws['G17'].border = Border(top = border_medium, bottom = border_medium)
	ws['H17'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)
	ws['G17'].value = "Loan Amount"
	ws.merge_cells('G18:H18')
	ws.merge_cells('G19:H19')
	ws.merge_cells('G20:H20')
	ws.merge_cells('G21:H21')
	ws.merge_cells('G22:H22')
	ws['G22'].border = Border(top = border_medium, bottom = border_medium)
	ws['H22'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)


	ws.merge_cells('I17:J17')
	ws['I17'].border = Border(top = border_medium, bottom = border_medium)
	ws['J17'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)
	ws['I17'].value = "Compensation"
	ws.merge_cells('I18:J18')
	ws.merge_cells('I19:J19')
	ws.merge_cells('I20:J20')
	ws.merge_cells('I21:J21')
	ws.merge_cells('I22:J22')
	ws['I22'].border = Border(top = border_medium, bottom = border_medium)
	ws['J22'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)


	ws['K17'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)
	ws['K17'].value = "CPL"
	ws['K22'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)


		#right box
	ws.merge_cells('M17:N17')
	ws['M17'].border = Border(top = border_medium, bottom = border_medium, left = border_medium)
	ws['N17'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)
	ws['M17'].value = "Quarter"
	ws.merge_cells('M18:N18')
	ws['M18'].border = Border(bottom = border_thin, left = border_medium)
	ws['N18'].border = Border(bottom = border_thin, right = border_medium)
	ws['M18'].value = current_year
	ws.merge_cells('M19:N19')
	ws['M19'].border = Border(bottom = border_thin, left = border_medium)
	ws['N19'].border = Border(bottom = border_thin, right = border_medium)
	ws['M19'].value = current_year - 1
	ws.merge_cells('M20:N20')
	ws['M20'].border = Border(bottom = border_thin, left = border_medium)
	ws['N20'].border = Border(bottom = border_thin, right = border_medium)
	ws['M20'].value = current_year - 2
	ws.merge_cells('M21:N21')
	ws['M21'].border = Border(bottom = border_thin, left = border_medium)
	ws['N21'].border = Border(bottom = border_thin, right = border_medium)
	ws['M21'].value = current_year - 3
	ws.merge_cells('M22:N22')
	ws['M22'].border = Border(left = border_medium)
	ws['N22'].border = Border(right = border_medium)
	ws['M22'].value = current_year - 4
	ws.merge_cells('M23:N23')
	ws['M23'].border = Border(top = border_medium, bottom = border_medium, left = border_medium)
	ws['N23'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)
	ws['M23'].value = "Grand Total"

	ws.merge_cells('O17:P17')
	ws['O17'].border = Border(top = border_medium, bottom = border_medium)
	ws['P17'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)
	ws['O17'].value = "Funded Loans"
	ws.merge_cells('O18:P18')
	ws.merge_cells('O19:P19')
	ws.merge_cells('O20:P20')
	ws.merge_cells('O21:P21')
	ws.merge_cells('O22:P22')
	ws.merge_cells('O23:P23')
	ws['O23'].border = Border(top = border_medium, bottom = border_medium)
	ws['P23'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)


	ws.merge_cells('Q17:R17')
	ws['Q17'].border = Border(top = border_medium, bottom = border_medium)
	ws['R17'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)
	ws['Q17'].value = "Loan Amount"
	ws.merge_cells('Q18:R18')
	ws.merge_cells('Q19:R19')
	ws.merge_cells('Q20:R20')
	ws.merge_cells('Q21:R21')
	ws.merge_cells('Q22:R22')
	ws.merge_cells('Q23:R23')
	ws['Q23'].border = Border(top = border_medium, bottom = border_medium)
	ws['R23'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)


	ws.merge_cells('S17:T17')
	ws['S17'].border = Border(top = border_medium, bottom = border_medium)
	ws['T17'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)
	ws['S17'].value = "Compensation"
	ws.merge_cells('S18:T18')
	ws.merge_cells('S19:T19')
	ws.merge_cells('S20:T20')
	ws.merge_cells('S21:T21')
	ws.merge_cells('S22:T22')
	ws.merge_cells('S23:T23')
	ws['S23'].border = Border(top = border_medium, bottom = border_medium)
	ws['T23'].border = Border(top = border_medium, bottom = border_medium, right = border_thin)


	ws['U17'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)
	ws['U17'].value = "CPL"
	ws['U23'].border = Border(top = border_medium, bottom = border_medium, right = border_medium)

		#inside borders left
	for i in range (18, 22):
		for j in range (5, 12):
			if i < 21 and (j == 5 or j == 7 or j == 9):
				ws.cell(row = i, column = j).border = Border(bottom = border_thin)
			elif i < 21 and (j == 6 or j == 8 or j == 10):
				ws.cell(row = i, column = j).border = Border(bottom = border_thin, right = border_thin)
			elif i < 21 and j == 11:
				ws.cell(row = i, column = j).border = Border(bottom = border_thin, right = border_medium)
			elif i == 21 and (j == 6 or j == 8 or j == 10):
				ws.cell(row = i, column = j).border = Border(right = border_thin)
			elif i == 21 and j == 11:
				ws.cell(row = i, column = j).border = Border(right = border_medium)


		#inside borders right
	for i in range (18, 23):
		for j in range (15, 22):
			if i < 22 and (j == 15 or j == 17 or j == 19):
				ws.cell(row = i, column = j).border = Border(bottom = border_thin)
			elif i < 22 and (j == 16 or j == 18 or j == 20):
				ws.cell(row = i, column = j).border = Border(bottom = border_thin, right = border_thin)
			elif i < 22 and j == 21:
				ws.cell(row = i, column = j).border = Border(bottom = border_thin, right = border_medium)
			elif i == 22 and (j == 16 or j == 18 or j == 20):
				ws.cell(row = i, column = j).border = Border(right = border_thin)
			elif i == 22 and j == 21:
				ws.cell(row = i, column = j).border = Border(right = border_medium)


		#alignment
	for i in range(17, 24):
		for j in range (3,22):
			ws.cell(row = i, column = j).alignment = align_center
			if i == 17 and j != 12:
				ws.cell(row = i, column = j).fill = fill_table_header
				ws.cell(row = i, column = j).font = font_table_headers
			if (i == 19 or i == 21) and j != 12:
				ws.cell(row = i, column = j).fill = fill_table_zebra
			if (i == 22 and j < 12) or (i == 23 and j > 12):
				ws.cell(row = i, column = j).fill = fill_table_total
				ws.cell(row = i, column = j).font = font_table_headers

		#number_formats
	for i in range (18, 24):
		for j in range (5, 22):
			if j == 5 or j == 15:
				ws.cell(row = i, column = j).number_format = "#,##0"
			if (j >= 7 and j<=10) or (j>=17 and j<=20):
				ws.cell(row = i, column = j).number_format = "$#,##0.00"
			if j == 11 or j == 21:
				ws.cell(row = i, column = j).number_format = "0.00%"



	#all borders, fills, font, formatting, and titles written, now writing data

	#calculates and fills the top monthly section
	monthly_funded, monthly_comp, monthly_cpl = calculate_monthly_values(affiliate_class, current_num_month, current_year)
	ws['D9'].value = monthly_funded
	ws['J9'].value = monthly_comp
	ws['P9'].value = monthly_cpl

	#calculates and fills the bottom left Quarter data
	current_year_data = calculate_quarterly_totals(affiliate_class, current_year)
	counter = 0
	for i in range(18, 23):
		ws.cell(row = i, column = 5).value = current_year_data[counter].total_loans
		ws.cell(row = i, column = 7).value = current_year_data[counter].total_funded_amt
		ws.cell(row = i, column = 9).value = current_year_data[counter].total_referral
		ws.cell(row = i, column = 11).value = current_year_data[counter].cpl
		counter += 1

	#calculates and fills the bottom right last 5 years data
	five_year_data = calculate_last_five_years_data(affiliate_class, current_year)
	counter = 0
	for i in range(18, 24):
		ws.cell(row = i, column = 15).value = five_year_data[counter].total_loans
		ws.cell(row = i, column = 17).value = five_year_data[counter].total_funded_amt
		ws.cell(row = i, column = 19).value = five_year_data[counter].total_referral
		ws.cell(row = i, column = 21).value = five_year_data[counter].cpl
		counter += 1


	wb.save(file)

def calculate_monthly_values(data, current_num_month, current_year):
	monthly_funded = 0
	monthly_comp = 0
	monthly_cpl = 0

	for i in data.record_data:
		if i.month == current_num_month and i.year == current_year:
			monthly_funded = i.loan_amt
			monthly_comp = i.referral_amt
			monthly_cpl = i.cpl

	return monthly_funded, monthly_comp, monthly_cpl

def calculate_quarterly_totals(data, current_year):
	qtr_array = ["Q1", "Q2", "Q3", "Q4", "Grand Total"]
	curr_year_data = [Output_Row(i) for i in qtr_array]
	for i in data.record_data:
		if i.year == current_year:
			if i.qtr == qtr_array[0]:
				curr_year_data[0].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
				curr_year_data[4].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			elif i.qtr == qtr_array[1]:
				curr_year_data[1].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
				curr_year_data[4].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			elif i.qtr == qtr_array[2]:
				curr_year_data[2].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
				curr_year_data[4].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			elif i.qtr == qtr_array[3]:
				curr_year_data[3].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
				curr_year_data[4].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)

	for i in curr_year_data:
		i.calc_cpl()
		i.format_empty_data()

	return curr_year_data

def calculate_last_five_years_data(data, current_year):
	years_array = []
	for i in range (0,5):
		years_array.append(current_year - i)
	years_array.append("Grand Total")

	five_year_data = [Output_Row(i) for i in years_array]
	for i in data.record_data:
		if i.year == years_array[0]:
			five_year_data[0].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			five_year_data[5].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
		elif i.year == years_array[1]:
			five_year_data[1].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			five_year_data[5].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
		elif i.year == years_array[2]:
			five_year_data[2].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			five_year_data[5].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
		elif i.year == years_array[3]:
			five_year_data[3].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			five_year_data[5].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
		elif i.year == years_array[4]:
			five_year_data[4].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)
			five_year_data[5].add_loans(i.num_loans_funded, i.loan_amt, i.referral_amt)

	for i in five_year_data:
		i.calc_cpl()
		i.format_empty_data()

	return five_year_data


""" End Output Generation Functions"""



""" ------------------------------------- End of Functions --------------------------------------------"""

""" ----------------------------------- Start of Code ---------------------------------------------"""
def main(current_month, current_year, input_file_path, output_file_path):


	#date dictionary to convert dates into the different formats required
	date_dicts = [
		{"numerical": "1", "short_date": "Jan", "long_date": "January"},
		{"numerical": "2", "short_date": "Feb", "long_date": "February"},
		{"numerical": "3", "short_date": "Mar", "long_date": "March"},
		{"numerical": "4", "short_date": "Apr", "long_date": "April"},
		{"numerical": "5", "short_date": "May", "long_date": "May"},
		{"numerical": "6", "short_date": "Jun", "long_date": "June"},
		{"numerical": "7", "short_date": "Jul", "long_date": "July"},
		{"numerical": "8", "short_date": "Aug", "long_date": "August"},
		{"numerical": "9", "short_date": "Sep", "long_date": "September"},
		{"numerical": "10", "short_date": "Oct", "long_date": "October"},
		{"numerical": "11", "short_date": "Nov", "long_date": "November"},
		{"numerical": "12", "short_date": "Dec", "long_date": "December"}
		]

	#loops thru the Date dictionary and sets the curr_short_month and curr_num_month variables based on inputs above
	for i in date_dicts:
		if current_month == i["long_date"]:
			current_short_month = i["short_date"]
			current_num_month = i["numerical"]
			current_num_month = int(current_num_month)

	affil_data = import_data(input_file_path)



	for i in affil_data:
		write_output_file(i, current_month, current_num_month, current_short_month, current_year, output_file_path)


""" ------------------------------------ End of Code ------------------------------------------------"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, Color, Fill, PatternFill
from sys import argv


user_name = "brian.sennott.marlette"
current_month = "February"
current_year = "2021"


current_year = int(current_year)



try:
	input_file_path = "C:/Users/" + user_name + "/Box/Finance Dept/FP&A/Month End Reports/"
	output_file_path = "C:/Users/" + user_name + "/Box/Finance Dept/FP&A/Month End Reports/Automated Affiliate Reports/"
	main(current_month, current_year, input_file_path, output_file_path)
except Exception as errormsg:
	print("error")
	print(errormsg)
	import traceback
	traceback.print_exc()
	input()
