#Pulls in the correct libraries
#for reading the csv and outputting the excel docs
from openpyxl import Workbook, load_workbook
import csv



#-----------------------------------Core functions-------------------------------------------

#this parses the CSV header and finds the corred header name, then returns the index position of that column
def parse_csv_header(row, search_string):
	for i, contents in enumerate(row):
		if row[i].lower() == search_string.lower():
			position = i

	return position

#this manages the format of the date string (converts it to the correct format if necessary)
def configure_date_formatting(orig_data):
	#checks to see if the data is in the correct format, if it is, just returns the original string, if not, formats the string properly
	if orig_data.find('/') == -1:
		new_data = orig_data
	else:
		#finds the index points for the first and second slashs, sets the day and year start indices
		first_slash = orig_data.find('/')
		day_start = first_slash + 1
		second_slash = orig_data.find('/',day_start)
		year_start = second_slash + 1

		#sets month based on above index positions, adds leading zeros if necessary
		if first_slash == 1:
			month = ("0" + orig_data[0:first_slash])
		elif first_slash == 2:
			month = (orig_data[0:first_slash])

		#sets day based on above index positions, adds leading zeros if necessary
		day = orig_data[day_start:second_slash]
		if len(day) == 1:
			day = "0" + day

		#sets year based on above index positions
		year = orig_data[year_start:]

		#returns a formatted concatonate based on year, month, and day
		new_data = f'{year}-{month}-{day}'

	return new_data


#this opens, parses the CSV, and adds the applicable columns and appends the result list
def read_and_total_affil_file(path, results_list, affil_count, dict, numerical_date):
	with open(path) as csv_file:
		csv_reader = csv.reader(csv_file, delimiter=',')
		line_count = 0
		funded_total = 0
		referral_total = 0
		total_funded_loans = 0

		for row in csv_reader:
			#For the first row in the file, it parses thru the header and finds the column number of each required column, then moves the line count down
			if line_count == 0:
				funded_position = parse_csv_header(row, dict["funded_string"])
				referral_position = parse_csv_header(row, dict["referral_string"])
				funded_date_position = parse_csv_header(row, dict["funded_month_string"])
				line_count +=  1


			else:
				#checks to see if there is any data in the first column of the row, if so it checks the date and verifies it falls within the correct month
				#then totals the funded and referral columns
				if row[1] != 0:
					#calls the configure date function to verify the date format, and reformat if necessary
					funded_date = configure_date_formatting(row[funded_date_position])

					if funded_date != "NULL" and funded_date[5:7] == numerical_date:
						if row[funded_position] != "NULL":
							funded_total += float(row[funded_position])
							total_funded_loans += 1
						else:
							funded_total += 0

						if row[referral_position] != "NULL":
							referral_total += float(row[referral_position])
						else:
							referral_total += 0

					line_count += 1

				else:
					funded_total = 0
					referral_total = 0
					total_funded_loans = 0



		#error handling in case there were no loans from the affiliate
		try:
			cost_per_loan = referral_total / funded_total
		except ZeroDivisionError:
			cost_per_loan = 0

		results_list.append({
			"count": affil_count,
			"name": dict["name"],
			"num_loans": total_funded_loans,
			"funded_total": funded_total,
			"referral_total": referral_total,
			"CPL": cost_per_loan,
			"file_path": path})
	return results_list, affil_count

#This takes the results list, opens a workbook, writes the header and then the results list to file
def write_output_file(results_list, save_path, save_file, affil_count):
	wb = Workbook()
	ws = wb.active
	ws.title = "Monthly output"
	ws['A1'] = "Affiliate Count"
	ws['B1'] = "Affiliate Name"
	ws['C1'] = "Funded Loans"
	ws['D1'] = "Funded Amount"
	ws['E1'] = "Referral Amount"
	ws['F1'] = "CPL"
	ws['G1'] = "Excel File Path"

	row_num = 2
	for r in results_list:
		col_num = 1
		for c in r.values():
			ws.cell(row = row_num, column = col_num).value = c
			col_num +=1
		row_num += 1

	#formats the final file
	for c in range(1, 8):
		for r in range(2, affil_count+1):
			v = ws.cell(row = r, column = c)
			if c == 3:
				v.number_format = "#,##0"
			elif c == 4 or c == 5:
				v.number_format = "#,##0.00"
			elif c == 6:
				v.number_format = "0.00%"

	wb.save(filename = save_path + save_file)

#This writes a file containing all of the errors that occurred during the script run
def write_error_file(error_list, save_path, error_file):
	file = open(save_path + error_file, "w")

	file.write("Monthly Check File Error List: \n\n")

	if not error_list:
		file.write("There were no errors running this script")
		print("There were no errors running this script")
	else:
		print("There were errors running this script, please consult the error log")

	for i in error_list:
		file.write(i)

	file.close()


#finds the last date file in the directory, solidifies that as the final file, appends the correct file name to the affil_dicts
def add_non_monthly_to_affil_dict(incomplete_dicts, complete_dicts, base_path, curr_year, curr_num_month, error_list):
	file_found = False

	for i in incomplete_dicts:
		for j in range(31, 9, -1):
			full_path = base_path+i["second_path"]+i["file_start"]+curr_year+"_"+curr_num_month+"_"+str(j)+".csv"
			try:
				with open(full_path) as csv_file:
					csv_reader = csv.reader(csv_file, delimiter=',')

				complete_dicts.append({
					"name": i["name"],
					"file_name": i["file_start"]+curr_year+"_"+curr_num_month+"_"+str(j)+".csv",
					"second_path": i["second_path"],
					"funded_string": i["funded_string"],
					"referral_string": i["referral_string"],
					"funded_month_string": i["funded_month_string"],
				})
				file_found = True
				break
			except:
				pass
		if not file_found:
			error_list.append(f'{i["name"]}: No file was found for this affiliate\n\n')

	return complete_dicts, error_list


"""-----------------------------Code Start--------------------------------------------------------"""
#---------------------------initial variable declarations----------------------------------------
wb = Workbook()
results_list = []
error_list = []


#This dictionary is used for outputs that do not have a file with a title of "monthly" and requires us to scan back thru the month to find the last date in the month file
non_monthly_file_dicts = [
	{"name": "Nerd Wallet", "file_start": "BestEgg_Nerd_Wallet_", "second_path": "Nerd Wallet Reporting/",
		"funded_string": "Funding_Amount", "referral_string": "NerdWallet_Commission", "funded_month_string": "Funding_Date" }
]

#starts the count of the rows to have a running count
affil_count = 1

#prompts user for their username, month, and year to insert into filedrive path and promts the user to verify
while True:
	user_name = input("enter your username: ")
	curr_month = input("enter the full name of the requested month: ")
	curr_month = curr_month.title()
	curr_year = input("enter the four digit year of the requested year: ")

	print("Your user name is " + user_name)
	print("The month selected is " + curr_month)
	print("The year selected is " + curr_year)
	print("Is this the correct username, month, and year? ")
	answer	= input("Y or N: ")
	answer = answer.upper()

	if answer == "Y":
		break

#debugging purposes only, remove at end
#user_name = 'chris.linden'
#curr_month = 'June'
#curr_year = '2020'



date_dicts = [
	{"numerical": "01", "short_date": "Jan", "long_date": "January"},
	{"numerical": "02", "short_date": "Feb", "long_date": "February"},
	{"numerical": "03", "short_date": "Mar", "long_date": "March"},
	{"numerical": "04", "short_date": "Apr", "long_date": "April"},
	{"numerical": "05", "short_date": "May", "long_date": "May"},
	{"numerical": "06", "short_date": "Jun", "long_date": "June"},
	{"numerical": "07", "short_date": "Jul", "long_date": "July"},
	{"numerical": "08", "short_date": "Aug", "long_date": "August"},
	{"numerical": "09", "short_date": "Sep", "long_date": "September"},
	{"numerical": "10", "short_date": "Oct", "long_date": "October"},
	{"numerical": "11", "short_date": "Nov", "long_date": "November"},
	{"numerical": "12", "short_date": "Dec", "long_date": "December"}
]


#loops thru the Date dictionary and sets the curr_short_month and curr_num_month variables based on inputs above
for i in date_dicts:
	if curr_month == i["long_date"]:
		curr_short_month = i["short_date"]
		curr_num_month = i["numerical"]

#affiliate dictionaries based on standard naming conventions
affiliate_dicts = [
	{ "name": "Credit Karma", "file_name": "CreditKarma_Marlette_v2_" + curr_month + curr_year + ".csv", "second_path": "Karma Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_Date" },
	{ "name": "Credit Sesame", "file_name": "Credit_Sesame_" + curr_month + "_" + curr_year + "_Funded_Loans.csv", "second_path": "Credit Sesame Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Even Financial", "file_name": "Even_Financial_Reporting_" + curr_month + "_" + curr_year + "_Funded.csv", "second_path": "Even Financial/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "QuinStreet", "file_name": "QuinStreet_" + curr_month + "_" + curr_year + "_Funded_loans.csv", "second_path": "QuinStreet Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Credit.com", "file_name": "Credit_com_Reporting_" + curr_month + "_" + curr_year + ".csv", "second_path": "Credit.com Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Magnify Money", "file_name": "Magnify_Money_" + curr_month + "_" + curr_year + "_Funded_Loans.csv", "second_path": "Magnify Money Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Money Tips", "file_name": "Money_Tips_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Money Tips/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Consumer Affairs", "file_name": "Consumer Affairs " + curr_month + " " + curr_year + " Funded Loans.csv", "second_path": "Consumer Affairs Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Buildzoom / Shogun", "file_name": "Shogun_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Shogun Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "ADV Market Direct", "file_name": "ADV_Market_Direct_" + curr_month + "_" + curr_year + "_Funded_Loans.csv", "second_path": "ADV Market Direct Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Leads Market", "file_name": "Leads_Market_Reporting_" + curr_month + "_" + curr_year + "_Funded.csv", "second_path": "Leads Market Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Best Company", "file_name": "Best_Company_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Best Company/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Credible", "file_name": "Credible_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Credible/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "LendEDU", "file_name": "LendEDU_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "LendEDU/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Finder.com", "file_name": "Finder_Com_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Finder.com/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "PrimeRates", "file_name": "PrimeRates_" + curr_month + "_" + curr_year + "_Funded_loans.csv", "second_path": "PrimeRates Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Invoice2Go", "file_name": "Invoice2go_" + curr_month + "_" + curr_year + "_Funded_Loans.csv", "second_path": "Invoice2go Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Super Money", "file_name":  "Super_Money_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "SuperMoney Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Monevo", "file_name": "Monevo_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Monevo Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "Make Lemonade", "file_name": "Make_Lemonade_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "Make Lemonade Reporting/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" },
	{ "name": "BankRate", "file_name": "BankRate_Funded_Loans_" + curr_month + "_" + curr_year + ".csv", "second_path": "BankRate Reporting/",
		"funded_string": "Funded_Loan_Amount", "referral_string": "Referral", "funded_month_string": "Funded_Date" },
	{ "name": "Experian", "file_name": "LC029_MONTHLY_REPORT_" + curr_year + curr_num_month + "01.csv", "second_path": "Experian/",
		"funded_string": "Funded_Amount", "referral_string": "Referral", "funded_month_string": "Funded_date" }
]

#sets the destination write path for the final excel file based on the user_name input
save_path = "C:/Users/" + user_name + "/Box/Finance Dept/FP&A/Month End Reports/Automated Affiliate Reports/"

#sets the users box path based on their username
base_path = "C:/Users/" + user_name + "/Box/"

#sets the save file
save_file = "Monthly Check File " + curr_short_month + " " + curr_year + ".xlsx"
error_file = "Monthly Check Error List " + curr_short_month + " " + curr_year + ".txt"

#runs thru non-monthly tallied files and builds the URL path and appends to the affiliate dictionary
affiliate_dicts, error_list = add_non_monthly_to_affil_dict(non_monthly_file_dicts, affiliate_dicts, base_path, curr_year, curr_num_month, error_list)

#parses thru each item in the affiliate dictionary list, building a results list for the writing function
for i in affiliate_dicts:
	full_path = base_path + i["second_path"] + i["file_name"]
	try:
		results_list, affil_count = read_and_total_affil_file(full_path, results_list, affil_count, i, curr_num_month)
	except FileNotFoundError as fnf:
		error_list.append(f'There was an issue with {i["name"]}\'s file path. \n {fnf} \n\n')
		affil_count -= 1
	except Exception as e:
		error_list.append(f'There was an issue with {i["name"]}. No other details are present. Please reach out to Chris Linden \n {e} \n\n')
		pass

	affil_count += 1

#calls the write output function
try:
	write_output_file(results_list, save_path, save_file, affil_count)
except PermissionError:
	print("\n\nThere is a permission error with the Output file, please make sure you do not have the Output file open before rerunning this script.\n\n")
	close_script = input("press enter to close this script")
	exit()

#calls the write error function
write_error_file(error_list, save_path, error_file)

#allows the user to just run the script external to the comamand prompt and gives them the time to process if there were or were not errors
close_script = input("press enter to close this script")
